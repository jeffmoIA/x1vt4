from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Pedido, ItemPedido
from .forms import PedidoForm
from carrito.models import Carrito
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Count
from django.db.models.functions import TruncMonth
import json
from django.utils import timezone
from datetime import datetime, timedelta
from django.core.paginator import Paginator
from django.views.generic import ListView
import csv
from django.http import HttpResponse
from datetime import datetime
from django.db.models import Q
from django.db.models import Sum, Avg, Count
import calendar
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from io import BytesIO
from django.contrib.auth.models import User
from django.template.loader import render_to_string
from django.http import JsonResponse
import json
from utils.logger import log_exception, log_audit, performance_monitor, exception_handler
from django.db.models import Prefetch
from utils.performance import query_debugger

@login_required
@performance_monitor(name="crear_pedido_view")  # Monitorea el tiempo de ejecución
@exception_handler(notify=True)  # Captura y registra excepciones
def crear_pedido(request):
    """Crear un nuevo pedido a partir del carrito del usuario"""
    # Obtener el carrito del usuario
    carrito, creado = Carrito.objects.get_or_create(usuario=request.user)
    
    # Verificar si el carrito está vacío
    if not carrito.items.exists():
        messages.warning(request, 'Tu carrito está vacío')
        return redirect('carrito:ver_carrito')
    
    if request.method == 'POST':
        form = PedidoForm(request.POST)
        if form.is_valid():
            # Crear un nuevo pedido pero no guardarlo todavía
            pedido = form.save(commit=False)
            pedido.usuario = request.user
            pedido.save()
            
            # Transferir items del carrito al pedido
            for item in carrito.items.all():
                ItemPedido.objects.create(
                    pedido=pedido,
                    producto=item.producto,
                    precio=item.producto.precio,
                    cantidad=item.cantidad
                )
            
            # Registrar acción de auditoría
            log_audit(
                user=request.user,
                action="create",
                object_type="Pedido",
                object_id=pedido.id,
                details={
                    "total_items": pedido.items.count(),
                    "total_amount": float(pedido.total())
                }
            )
            
            # Vaciar el carrito
            carrito.items.all().delete()
            
            # Mostrar mensaje de éxito
            messages.success(request, 'Pedido realizado con éxito')
            
            # Redirigir a la página de detalles del pedido
            return redirect('pagos:seleccionar_metodo_pago', pedido_id=pedido.id)

    else:
        # Prellenar el formulario con la información del usuario si existe
        initial_data = {}
        if hasattr(request.user, 'perfil'):
            perfil = request.user.perfil
            initial_data = {
                'nombre_completo': request.user.get_full_name() or request.user.username,
                'direccion': perfil.direccion or '',
                'ciudad': perfil.ciudad or '',
                'codigo_postal': perfil.codigo_postal or '',
                'telefono': perfil.telefono or '',
            }
        form = PedidoForm(initial=initial_data)
    
    return render(request, 'pedidos/crear_pedido.html', {
        'form': form,
        'carrito': carrito
    })

@login_required
def lista_pedidos(request):
    """Mostrar todos los pedidos del usuario"""
    pedidos = Pedido.objects.filter(usuario=request.user).order_by('-fecha_pedido')
    return render(request, 'pedidos/lista_pedidos.html', {
        'pedidos': pedidos
    })

@login_required
def detalle_pedido(request, pedido_id):
    """Mostrar los detalles de un pedido específico"""
    # Usamos prefetch_related para cargar los items relacionados y sus productos en una sola consulta
    pedido = get_object_or_404(
        Pedido.objects.prefetch_related('items__producto'),
        id=pedido_id, 
        usuario=request.user
    )
    return render(request, 'pedidos/detalle_pedido.html', {
        'pedido': pedido
    })

@login_required
def cancelar_pedido(request, pedido_id):
    """Cancelar un pedido pendiente"""
    pedido = get_object_or_404(Pedido, id=pedido_id, usuario=request.user)
    
    # Solo se pueden cancelar pedidos pendientes
    if pedido.estado == 'pendiente':
        if request.method == 'POST':
            pedido.estado = 'cancelado'
            pedido.save()
            messages.success(request, 'Pedido cancelado exitosamente')
            return redirect('pedidos:lista_pedidos')
        
        return render(request, 'pedidos/confirmar_cancelacion.html', {
            'pedido': pedido
        })
    else:
        messages.error(request, 'Este pedido no puede ser cancelado')
        return redirect('pedidos:detalle_pedido', pedido_id=pedido.id)
    
def obtener_factura(request, pedido_id):
    """Vista para descargar la factura de un pedido"""
    # Verificar que el usuario esté autenticado
    if not request.user.is_authenticated:
        return redirect('usuarios:login')
    
    # Verificar que el pedido pertenece al usuario
    pedido = get_object_or_404(Pedido, id=pedido_id, usuario=request.user)
    
    # Importar la función desde utils
    from .utils import obtener_factura
    return obtener_factura(request, pedido_id)

# Función para verificar si el usuario es administrador
def es_admin(user):
    return user.is_staff

@login_required
@user_passes_test(es_admin)  # Solo administradores pueden acceder
def estadisticas_pedidos(request):
    """
    Vista para mostrar estadísticas de pedidos para administradores
    """
    # Obtener todos los pedidos
    pedidos = Pedido.objects.all()
    
    # Estadísticas generales
    total_pedidos = pedidos.count()
    pedidos_pendientes = pedidos.filter(estado=Pedido.PENDIENTE).count()
    pedidos_procesando = pedidos.filter(estado=Pedido.PROCESANDO).count()
    pedidos_enviados = pedidos.filter(estado=Pedido.ENVIADO).count()
    pedidos_entregados = pedidos.filter(estado=Pedido.ENTREGADO).count()
    pedidos_cancelados = pedidos.filter(estado=Pedido.CANCELADO).count()
    
    # Obtener pedidos por mes (últimos 6 meses)
    fecha_inicio = timezone.now() - timedelta(days=180)  # 6 meses atrás
    pedidos_por_mes_query = pedidos.filter(
        fecha_pedido__gte=fecha_inicio
    ).annotate(
        mes=TruncMonth('fecha_pedido')
    ).values('mes').annotate(
        total=Count('id')
    ).order_by('mes')
    
    # Preparar datos para el gráfico de pedidos por mes
    meses = []
    totales = []
    
    for item in pedidos_por_mes_query:
        meses.append(item['mes'].strftime('%b %Y'))
        totales.append(item['total'])
    
    # Si no hay datos para algunos meses, llenar con ceros
    # Esto asegura que siempre mostramos 6 meses incluso si no hay pedidos
    if len(meses) < 6:
        fecha_actual = timezone.now()
        for i in range(6):
            fecha = fecha_actual - timedelta(days=30 * i)
            mes_str = fecha.strftime('%b %Y')
            if mes_str not in meses:
                meses.insert(0, mes_str)
                totales.insert(0, 0)
        # Limitar a solo 6 meses
        meses = meses[:6]
        totales = totales[:6]
    
    # Ordenar cronológicamente
    meses.reverse()
    totales.reverse()
    
    # Obtener pedidos recientes (últimos 10)
    pedidos_recientes = pedidos.order_by('-fecha_pedido')[:10]
    
    # Contexto para la plantilla
    context = {
        'total_pedidos': total_pedidos,
        'pedidos_pendientes': pedidos_pendientes,
        'pedidos_procesando': pedidos_procesando,
        'pedidos_enviados': pedidos_enviados,
        'pedidos_entregados': pedidos_entregados,
        'pedidos_cancelados': pedidos_cancelados,
        'meses_labels': json.dumps(meses),
        'pedidos_por_mes': totales,
        'pedidos_recientes': pedidos_recientes,
    }
    
    # Métricas adicionales
    valor_promedio_pedidos = pedidos.aggregate(avg=Avg('items__precio'))['avg'] or 0
    
    # Ventas totales
    ventas_totales = sum(pedido.total() for pedido in pedidos)
    
    # Ventas por mes (último año)
    fecha_inicio_anual = timezone.now() - timedelta(days=365)
    
    # Mejores clientes
    mejores_clientes = User.objects.annotate(
        total_pedidos=Count('pedidos'),
        gasto_total=Sum('pedidos__items__precio')
    ).order_by('-gasto_total')[:5]
    
    # Añadir al contexto
    context.update({
        'valor_promedio_pedidos': valor_promedio_pedidos,
        'ventas_totales': ventas_totales,
        'mejores_clientes': mejores_clientes,
    })
    
    return render(request, 'pedidos/admin/estadisticas.html', context)

@login_required
@user_passes_test(es_admin)
@query_debugger
def admin_lista_pedidos(request):
    """Vista optimizada para gestión de pedidos en el panel de administración"""
    # Usamos select_related para cargar relaciones en una sola consulta
    # y prefetch_related para optimizar la carga de items
    pedidos = (Pedido.objects
              .select_related('usuario')
              .prefetch_related(
                  Prefetch('items', 
                           queryset=ItemPedido.objects.select_related('producto')
                  ),
                  'historial_estados'
              ))
              
    # Añadimos paginación eficiente para manejar grandes volúmenes
    paginator = Paginator(pedidos.order_by('-fecha_pedido'), 50)  # 50 pedidos por página
    page = request.GET.get('page', 1)
    
    try:
        pedidos_paginados = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        pedidos_paginados = paginator.page(1)
    
    context = {
        'pedidos': pedidos_paginados,
    }
    
    return render(request, 'pedidos/admin/lista_pedidos.html', context)

@login_required
@user_passes_test(es_admin)
def cambiar_estado_pedido(request, pedido_id):
    """Vista para que los administradores cambien el estado de un pedido"""
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('nuevo_estado')
        if nuevo_estado in dict(Pedido.ESTADOS).keys():
            notas = f"Cambio realizado por el administrador {request.user.username}"
            pedido.cambiar_estado(nuevo_estado, notas)
            
            # Si es una solicitud AJAX, devolver respuesta JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'mensaje': f'Estado del pedido #{pedido.id} actualizado a {nuevo_estado}'
                })
            
            # Para solicitudes normales
            messages.success(request, f'Estado del pedido #{pedido.id} actualizado a {dict(Pedido.ESTADOS)[nuevo_estado]}')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': 'Estado no válido'
                }, status=400)
            
            messages.error(request, 'Estado no válido')
    
    # Redirigir a la lista de pedidos
    referer = request.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    else:
        return redirect('pedidos:admin_lista_pedidos')
    
@login_required
@user_passes_test(es_admin)
def admin_detalle_pedido(request, pedido_id):
    """Vista detallada de pedido para administradores"""
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    # Obtener el historial de estados del pedido
    historial = pedido.historial_estados.all().order_by('-fecha_cambio')
    
    context = {
        'pedido': pedido,
        'historial': historial,
    }
    
    return render(request, 'pedidos/admin/detalle_pedido.html', context)

@login_required
@user_passes_test(es_admin)
def actualizar_seguimiento(request, pedido_id):
    """Actualizar información de seguimiento de un pedido"""
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    if request.method == 'POST':
        # Actualizar los campos de seguimiento
        pedido.empresa_envio = request.POST.get('empresa_envio', '')
        pedido.codigo_seguimiento = request.POST.get('codigo_seguimiento', '')
        pedido.save()
        
        messages.success(request, f'Información de seguimiento del pedido #{pedido.id} actualizada')
    
    return redirect('pedidos:admin_detalle_pedido', pedido_id=pedido.id)
    

@login_required
@user_passes_test(es_admin)
def exportar_pedidos_excel(request):
    """Exportar pedidos a Excel con formato adecuado"""
    # Aplicar filtros (similar a la vista de lista)
    queryset = Pedido.objects.all().order_by('-fecha_pedido')
    estado = request.GET.get('estado')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if estado:
        queryset = queryset.filter(estado=estado)
    if fecha_desde:
        queryset = queryset.filter(fecha_pedido__date__gte=fecha_desde)
    if fecha_hasta:
        queryset = queryset.filter(fecha_pedido__date__lte=fecha_hasta)
    
    # Crear un libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos"
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    centered_alignment = Alignment(horizontal='center')
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    
    # Encabezados
    columns = ['ID', 'Cliente', 'Email', 'Fecha', 'Estado', 'Total', 'Productos']
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered_alignment
        cell.border = border
    
    # Agregar datos
    for row_num, pedido in enumerate(queryset, 2):
        # Lista de productos formateada
        productos = ", ".join([f"{item.cantidad}x {item.producto.nombre}" for item in pedido.items.all()])
        
        # Datos de la fila
        row = [
            pedido.id, 
            pedido.nombre_completo,
            pedido.usuario.email,
            pedido.fecha_pedido.strftime("%d/%m/%Y %H:%M"),
            dict(Pedido.ESTADOS)[pedido.estado],
            f"${pedido.total()}",
            productos
        ]
        
        # Escribir la fila y aplicar estilo
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = border
            
            # Centrar algunas columnas
            if col_num in [1, 4, 5]:  # ID, Fecha, Estado
                cell.alignment = centered_alignment
    
    # Autoajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="pedidos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Guardar el libro de trabajo en el objeto de respuesta
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())
    
    return response

@login_required
@user_passes_test(es_admin)
def cambiar_estado_pedido_ajax(request, pedido_id):
    """Vista AJAX para cambiar el estado de un pedido"""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        pedido = get_object_or_404(Pedido, id=pedido_id)
        nuevo_estado = request.POST.get('nuevo_estado')
        
        if nuevo_estado in dict(Pedido.ESTADOS).keys():
            notas = f"Cambio realizado por {request.user.username} vía AJAX"
            pedido.cambiar_estado(nuevo_estado, notas)
            
            return JsonResponse({
                'success': True,
                'estado': nuevo_estado,
                'estado_display': dict(Pedido.ESTADOS)[nuevo_estado]
            })
        
        return JsonResponse({'success': False, 'error': 'Estado no válido'}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

